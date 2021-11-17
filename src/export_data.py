import os
import re
import csv
import time
import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook

EXCEL_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                      'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def clean_text(string):
    pattern_new_line = r'\n'
    clean_n_result = re.sub(pattern_new_line, '', string)

    pattern_multi_spaces = r'\s{2,}'
    multi_spaces_result = re.sub(pattern_multi_spaces, ' ', clean_n_result)

    pattern_head_tail_spaces = r'^\s+|\s+$'
    result = re.sub(pattern_head_tail_spaces, '', multi_spaces_result)
    return result


def real_time(elapsed=00000):
    # show real time
    if elapsed < 0:
        raise ValueError('Elapsed negative')
    hours, rem = divmod(elapsed, 3600)
    minutes, seconds = divmod(rem, 60)
    result = '{:0>2}:{:0>2}:{:05.2f}'.format(int(hours), int(minutes), seconds)
    return result


def save_csv(list_data=[], save_type=None, header=['H1', 'H2', 'H3'], file_name='export_data'):
    with open(f'{file_name}.csv', 'w', encoding='utf-8', newline='', ) as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(list_data)


def save_excel(list_data=[], save_type=None, header=['H1', 'H2', 'H3'], file_name='export_data'):
    # create excel file
    wb = Workbook()
    wb1 = wb.active
    wb1.title = "Sheet1"
    wb.save(f'{file_name}.xlsx')

    # open file
    workbook = load_workbook(filename=f'{file_name}.xlsx')
    sheet_file = workbook['Sheet1']
    for i, head in enumerate(header):
        sheet_file[EXCEL_COLUMNS[i]+'1'] = head

    if len(list_data) >= 0:
        for i, row in enumerate(list_data):
            for j, head in enumerate(header):
                sheet_file[EXCEL_COLUMNS[j]+str(i+2)] = row[j]
    workbook.save(filename=f'{file_name}.xlsx')


def csv2excel(csv_filename='sample_output.csv', excel_filename='sample_output_converted.xlsx'):
    if os.path.isfile(csv_filename):
        df = pd.read_csv(csv_filename)
        df.to_excel(excel_filename, index=False)
    else:
        print(f'No such file or directory: \'{csv_filename}\'')


def save_data(list_data=[], save_type=None, header=['H1', 'H2', 'H3'], file_name='export_data'):
    # save to csv file
    if save_type == 'csv':
        save_csv(list_data=list_data, header=header, file_name=file_name)
    elif save_type == 'excel':
        save_excel(list_data=list_data, header=header, file_name=file_name)
    elif save_type == 'all':
        save_csv(list_data=list_data, header=header, file_name=file_name)
        save_excel(list_data=list_data, header=header, file_name=file_name)


if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description='export data', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-s', '--save', type=str,
                        choices=['csv', 'excel', 'all'], default=None, help='save as csv or excel file')
    args = parser.parse_args()
    save_type = args.save
    start_time = time.time()
    # sample output
    csv2excel('list_steam_games.csv')
    save_data(list_data=[(1, 2, 4, 5)], save_type='excel',
              file_name='sample_output')
    save_data(list_data=[(1, 2, 4, 5)], save_type=save_type)
    end_time = time.time()
    print('Elapsed:', real_time(end_time-start_time))
